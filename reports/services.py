from django.db.models import Q, Count, Avg, Sum
from django.utils import timezone
from django.core.cache import cache
from django.contrib.auth import get_user_model
from typing import Dict, Any, List, Optional
import io
import hashlib
import json
from datetime import datetime, timedelta

# Optional imports - will be used when libraries are installed
try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False

try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.chart import BarChart, Reference
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from .models import ReportExecution, ReportTemplate

User = get_user_model()


class CacheService:
    """Service for managing report caching"""

    def generate_cache_key(self, report_type: str, filters: Dict[str, Any], user_id: int) -> str:
        """Generate a unique cache key for report data"""
        # Create a hash of the filters for uniqueness
        filters_str = json.dumps(filters, sort_keys=True, default=str)
        filters_hash = hashlib.md5(filters_str.encode()).hexdigest()[:8]

        return f"report_{report_type}_{user_id}_{filters_hash}"

    def get_cached_report(self, cache_key: str) -> Optional[Dict[str, Any]]:
        """Get cached report data"""
        return cache.get(cache_key)

    def cache_report(self, cache_key: str, data: Dict[str, Any], timeout: int = 300) -> None:
        """Cache report data"""
        cache.set(cache_key, data, timeout)

    def invalidate_report_cache(self, report_type: str) -> None:
        """Invalidate all cached reports of a specific type"""
        # In a real implementation, you might use cache versioning or tags
        # For now, we'll rely on TTL
        pass


class ReportService:
    """Service for generating reports"""

    def execute_report(self, execution: ReportExecution) -> Dict[str, Any]:
        """Execute a report and return results"""
        execution.start_execution()

        try:
            # Generate report based on template type
            filters = execution.parameters
            data = self.generate_report(
                execution.template.report_type,
                filters,
                execution.executed_by
            )

            # Save results
            execution.complete_execution(
                result_data=data,
                rows_processed=len(data.get('data', []))
            )

            return data

        except Exception as e:
            execution.fail_execution(str(e))
            raise

    def generate_report(self, report_type: str, filters: Dict[str, Any], user: User) -> Dict[str, Any]:
        """Generate report data based on type"""
        if report_type == 'employees':
            return self._generate_employees_report(filters, user)
        elif report_type == 'terminations':
            return self._generate_terminations_report(filters, user)
        elif report_type == 'evaluations':
            return self._generate_evaluations_report(filters, user)
        elif report_type == 'leave_requests':
            return self._generate_leave_requests_report(filters, user)
        elif report_type == 'admissions':
            return self._generate_admissions_report(filters, user)
        else:
            raise ValueError(f"Unknown report type: {report_type}")

    def _generate_employees_report(self, filters: Dict[str, Any], user: User) -> Dict[str, Any]:
        """Generate employees report"""
        try:
            from employees.models import Employee

            queryset = Employee.objects.all()

            # Apply filters
            if filters.get('department'):
                queryset = queryset.filter(department__icontains=filters['department'])

            if filters.get('position'):
                queryset = queryset.filter(position__icontains=filters['position'])

            if filters.get('employee_status'):
                queryset = queryset.filter(status=filters['employee_status'])

            if filters.get('start_date'):
                queryset = queryset.filter(hire_date__gte=filters['start_date'])

            if filters.get('end_date'):
                queryset = queryset.filter(hire_date__lte=filters['end_date'])

            # Ordering
            ordering = filters.get('ordering', '-hire_date')
            queryset = queryset.order_by(ordering)

            # Convert to list
            employees_data = []
            for emp in queryset:
                employees_data.append({
                    'id': emp.id,
                    'employee_id': emp.employee_id,
                    'full_name': emp.full_name,
                    'email': emp.email,
                    'department': emp.department,
                    'position': emp.position,
                    'hire_date': emp.hire_date.isoformat() if emp.hire_date else None,
                    'salary': float(emp.salary) if emp.salary else None,
                    'status': emp.get_status_display(),
                    'phone': emp.phone,
                    'education_level': emp.get_education_level_display() if emp.education_level else None,
                })

            # Generate summary
            total_count = queryset.count()
            avg_salary = queryset.aggregate(avg_salary=Avg('salary'))['avg_salary'] or 0

            # Department breakdown
            dept_breakdown = queryset.values('department').annotate(
                count=Count('id')
            ).order_by('-count')

            # Status breakdown
            status_breakdown = queryset.values('status').annotate(
                count=Count('id')
            )

            return {
                'report_type': 'employees',
                'generated_at': timezone.now().isoformat(),
                'filters': filters,
                'summary': {
                    'total_employees': total_count,
                    'average_salary': float(avg_salary),
                    'department_breakdown': list(dept_breakdown),
                    'status_breakdown': list(status_breakdown),
                },
                'data': employees_data,
                'total_records': total_count
            }

        except Exception as e:
            raise Exception(f"Error generating employees report: {str(e)}")

    def _generate_terminations_report(self, filters: Dict[str, Any], user: User) -> Dict[str, Any]:
        """Generate terminations report"""
        try:
            from termination.models import TerminationRequest

            queryset = TerminationRequest.objects.select_related(
                'funcionario', 'motivo', 'solicitante'
            )

            # Apply filters
            if filters.get('start_date'):
                queryset = queryset.filter(data_desligamento__gte=filters['start_date'])

            if filters.get('end_date'):
                queryset = queryset.filter(data_desligamento__lte=filters['end_date'])

            # Ordering
            ordering = filters.get('ordering', '-created_at')
            queryset = queryset.order_by(ordering)

            # Convert to list
            terminations_data = []
            for term in queryset:
                terminations_data.append({
                    'id': term.id,
                    'employee_name': term.funcionario.get_full_name(),
                    'employee_email': term.funcionario.email,
                    'reason': term.motivo.nome,
                    'termination_date': term.data_desligamento.isoformat(),
                    'last_work_day': term.data_ultimo_dia.isoformat(),
                    'status': term.get_status_display(),
                    'requester': term.solicitante.get_full_name(),
                    'justification': term.justificativa,
                    'created_at': term.created_at.isoformat(),
                })

            # Generate summary
            total_count = queryset.count()

            # Reason breakdown
            reason_breakdown = queryset.values('motivo__nome').annotate(
                count=Count('id')
            ).order_by('-count')

            # Status breakdown
            status_breakdown = queryset.values('status').annotate(
                count=Count('id')
            )

            # Monthly breakdown (last 12 months)
            monthly_breakdown = []
            for i in range(12):
                month_start = timezone.now().replace(day=1) - timedelta(days=30*i)
                month_end = month_start + timedelta(days=31)
                count = queryset.filter(
                    data_desligamento__gte=month_start,
                    data_desligamento__lt=month_end
                ).count()
                monthly_breakdown.append({
                    'month': month_start.strftime('%Y-%m'),
                    'count': count
                })

            return {
                'report_type': 'terminations',
                'generated_at': timezone.now().isoformat(),
                'filters': filters,
                'summary': {
                    'total_terminations': total_count,
                    'reason_breakdown': list(reason_breakdown),
                    'status_breakdown': list(status_breakdown),
                    'monthly_breakdown': monthly_breakdown,
                },
                'data': terminations_data,
                'total_records': total_count
            }

        except Exception as e:
            raise Exception(f"Error generating terminations report: {str(e)}")

    def _generate_evaluations_report(self, filters: Dict[str, Any], user: User) -> Dict[str, Any]:
        """Generate evaluations report"""
        try:
            from evaluations.models import PerformanceEvaluation

            queryset = PerformanceEvaluation.objects.select_related(
                'employee', 'evaluator'
            )

            # Apply filters
            if filters.get('start_date'):
                queryset = queryset.filter(evaluation_date__gte=filters['start_date'])

            if filters.get('end_date'):
                queryset = queryset.filter(evaluation_date__lte=filters['end_date'])

            # Ordering
            ordering = filters.get('ordering', '-evaluation_date')
            queryset = queryset.order_by(ordering)

            # Convert to list
            evaluations_data = []
            for eval in queryset:
                evaluations_data.append({
                    'id': eval.id,
                    'employee_name': eval.employee.get_full_name(),
                    'employee_department': getattr(eval.employee, 'department', ''),
                    'evaluator_name': eval.evaluator.get_full_name(),
                    'evaluation_date': eval.evaluation_date.isoformat(),
                    'period': f"{eval.period_start.isoformat()} to {eval.period_end.isoformat()}",
                    'overall_score': float(eval.overall_score) if eval.overall_score else None,
                    'status': eval.get_status_display(),
                    'created_at': eval.created_at.isoformat(),
                })

            # Generate summary
            total_count = queryset.count()
            avg_score = queryset.aggregate(avg_score=Avg('overall_score'))['avg_score'] or 0

            # Status breakdown
            status_breakdown = queryset.values('status').annotate(
                count=Count('id')
            )

            # Score distribution
            score_ranges = [
                ('Excelente (4.5-5.0)', queryset.filter(overall_score__gte=4.5).count()),
                ('Bom (3.5-4.4)', queryset.filter(overall_score__gte=3.5, overall_score__lt=4.5).count()),
                ('Regular (2.5-3.4)', queryset.filter(overall_score__gte=2.5, overall_score__lt=3.5).count()),
                ('Ruim (1.0-2.4)', queryset.filter(overall_score__gte=1.0, overall_score__lt=2.5).count()),
            ]

            return {
                'report_type': 'evaluations',
                'generated_at': timezone.now().isoformat(),
                'filters': filters,
                'summary': {
                    'total_evaluations': total_count,
                    'average_score': float(avg_score),
                    'status_breakdown': list(status_breakdown),
                    'score_distribution': score_ranges,
                },
                'data': evaluations_data,
                'total_records': total_count
            }

        except Exception as e:
            raise Exception(f"Error generating evaluations report: {str(e)}")

    def _generate_leave_requests_report(self, filters: Dict[str, Any], user: User) -> Dict[str, Any]:
        """Generate leave requests report"""
        try:
            from leave_requests.models import LeaveRequest

            queryset = LeaveRequest.objects.select_related('employee')

            # Apply filters
            if filters.get('start_date'):
                queryset = queryset.filter(start_date__gte=filters['start_date'])

            if filters.get('end_date'):
                queryset = queryset.filter(end_date__lte=filters['end_date'])

            # Ordering
            ordering = filters.get('ordering', '-created_at')
            queryset = queryset.order_by(ordering)

            # Convert to list
            leave_requests_data = []
            for leave in queryset:
                leave_requests_data.append({
                    'id': leave.id,
                    'employee_name': leave.employee.get_full_name(),
                    'employee_department': getattr(leave.employee, 'department', ''),
                    'leave_type': leave.get_leave_type_display(),
                    'start_date': leave.start_date.isoformat(),
                    'end_date': leave.end_date.isoformat(),
                    'days_requested': leave.days_requested,
                    'status': leave.get_status_display(),
                    'reason': leave.reason,
                    'created_at': leave.created_at.isoformat(),
                })

            # Generate summary
            total_count = queryset.count()
            total_days = queryset.aggregate(total_days=Sum('days_requested'))['total_days'] or 0

            # Status breakdown
            status_breakdown = queryset.values('status').annotate(
                count=Count('id')
            )

            # Type breakdown
            type_breakdown = queryset.values('leave_type').annotate(
                count=Count('id')
            ).order_by('-count')

            return {
                'report_type': 'leave_requests',
                'generated_at': timezone.now().isoformat(),
                'filters': filters,
                'summary': {
                    'total_requests': total_count,
                    'total_days_requested': total_days,
                    'status_breakdown': list(status_breakdown),
                    'type_breakdown': list(type_breakdown),
                },
                'data': leave_requests_data,
                'total_records': total_count
            }

        except Exception as e:
            raise Exception(f"Error generating leave requests report: {str(e)}")

    def _generate_admissions_report(self, filters: Dict[str, Any], user: User) -> Dict[str, Any]:
        """Generate admissions report"""
        try:
            from employees.models import Employee, AdmissionProcess

            queryset = Employee.objects.filter(
                admission_process__isnull=False
            ).select_related('admission_process')

            # Apply filters
            if filters.get('start_date'):
                queryset = queryset.filter(hire_date__gte=filters['start_date'])

            if filters.get('end_date'):
                queryset = queryset.filter(hire_date__lte=filters['end_date'])

            if filters.get('employee_status'):
                queryset = queryset.filter(status=filters['employee_status'])

            # Ordering
            ordering = filters.get('ordering', '-hire_date')
            queryset = queryset.order_by(ordering)

            # Convert to list
            admissions_data = []
            for emp in queryset:
                process = emp.admission_process
                admissions_data.append({
                    'id': emp.id,
                    'employee_id': emp.employee_id,
                    'full_name': emp.full_name,
                    'email': emp.email,
                    'department': emp.department,
                    'position': emp.position,
                    'hire_date': emp.hire_date.isoformat() if emp.hire_date else None,
                    'admission_status': process.get_status_display(),
                    'completion_percentage': process.completion_percentage,
                    'personal_info_completed': process.personal_info_completed,
                    'documents_uploaded': process.documents_uploaded,
                    'hr_review_completed': process.hr_review_completed,
                    'process_started': process.started_at.isoformat(),
                    'process_completed': process.completed_at.isoformat() if process.completed_at else None,
                })

            # Generate summary
            total_count = queryset.count()

            # Status breakdown
            status_breakdown = AdmissionProcess.objects.values('status').annotate(
                count=Count('id')
            )

            # Completion stages
            completion_stats = {
                'personal_info_completed': AdmissionProcess.objects.filter(personal_info_completed=True).count(),
                'documents_uploaded': AdmissionProcess.objects.filter(documents_uploaded=True).count(),
                'hr_review_completed': AdmissionProcess.objects.filter(hr_review_completed=True).count(),
            }

            return {
                'report_type': 'admissions',
                'generated_at': timezone.now().isoformat(),
                'filters': filters,
                'summary': {
                    'total_admissions': total_count,
                    'status_breakdown': list(status_breakdown),
                    'completion_stats': completion_stats,
                },
                'data': admissions_data,
                'total_records': total_count
            }

        except Exception as e:
            raise Exception(f"Error generating admissions report: {str(e)}")


class ExportService:
    """Service for exporting reports to different formats"""

    def export_to_pdf(self, data: Dict[str, Any], report_type: str, include_charts: bool = False) -> bytes:
        """Export report data to PDF"""
        if not REPORTLAB_AVAILABLE:
            raise Exception("ReportLab library is not installed. Please install it to export PDF reports.")

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []

        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=1  # Center
        )

        # Title
        title = f"Relatório de {report_type.replace('_', ' ').title()}"
        elements.append(Paragraph(title, title_style))
        elements.append(Spacer(1, 20))

        # Summary section
        if 'summary' in data:
            summary_title = Paragraph("Resumo", styles['Heading2'])
            elements.append(summary_title)
            elements.append(Spacer(1, 10))

            summary = data['summary']
            summary_data = []

            for key, value in summary.items():
                if isinstance(value, (int, float)):
                    summary_data.append([key.replace('_', ' ').title(), str(value)])
                elif isinstance(value, list) and len(value) <= 10:  # Limit list items
                    items = ', '.join([str(item) for item in value[:5]])
                    if len(value) > 5:
                        items += '...'
                    summary_data.append([key.replace('_', ' ').title(), items])

            if summary_data:
                summary_table = Table(summary_data)
                summary_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 12),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                elements.append(summary_table)
                elements.append(Spacer(1, 20))

        # Data table
        if 'data' in data and data['data']:
            data_title = Paragraph("Dados Detalhados", styles['Heading2'])
            elements.append(data_title)
            elements.append(Spacer(1, 10))

            # Get first few records for display
            records = data['data'][:50]  # Limit to 50 records for PDF

            if records:
                # Get headers from first record
                headers = list(records[0].keys())
                table_data = [headers]

                # Add data rows
                for record in records:
                    row = []
                    for header in headers:
                        value = record.get(header, '')
                        if value is None:
                            value = ''
                        # Truncate long text
                        str_value = str(value)
                        if len(str_value) > 30:
                            str_value = str_value[:27] + '...'
                        row.append(str_value)
                    table_data.append(row)

                # Create table
                table = Table(table_data)
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 8),
                    ('FONTSIZE', (0, 1), (-1, -1), 7),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                elements.append(table)

                if len(data['data']) > 50:
                    note = Paragraph(
                        f"Nota: Exibindo primeiros 50 registros de {len(data['data'])} total.",
                        styles['Normal']
                    )
                    elements.append(Spacer(1, 10))
                    elements.append(note)

        # Generation info
        gen_info = Paragraph(
            f"Relatório gerado em: {data.get('generated_at', timezone.now().isoformat())}",
            styles['Normal']
        )
        elements.append(Spacer(1, 20))
        elements.append(gen_info)

        doc.build(elements)
        buffer.seek(0)
        return buffer.getvalue()

    def export_to_excel(self, data: Dict[str, Any], report_type: str) -> bytes:
        """Export report data to Excel"""
        if not OPENPYXL_AVAILABLE:
            raise Exception("OpenPyXL library is not installed. Please install it to export Excel reports.")

        wb = Workbook()
        ws = wb.active
        ws.title = f"{report_type.replace('_', ' ').title()} Report"

        # Header style
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Title
        ws['A1'] = f"Relatório de {report_type.replace('_', ' ').title()}"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:F1')

        # Generation date
        ws['A2'] = f"Gerado em: {data.get('generated_at', timezone.now().isoformat())}"
        ws.merge_cells('A2:F2')

        current_row = 4

        # Summary section
        if 'summary' in data:
            ws[f'A{current_row}'] = "RESUMO"
            ws[f'A{current_row}'].font = Font(bold=True, size=14)
            current_row += 2

            summary = data['summary']
            for key, value in summary.items():
                if isinstance(value, (int, float, str)):
                    ws[f'A{current_row}'] = key.replace('_', ' ').title()
                    ws[f'B{current_row}'] = value
                    current_row += 1

            current_row += 2

        # Data section
        if 'data' in data and data['data']:
            ws[f'A{current_row}'] = "DADOS DETALHADOS"
            ws[f'A{current_row}'].font = Font(bold=True, size=14)
            current_row += 2

            records = data['data']
            if records:
                # Headers
                headers = list(records[0].keys())
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=current_row, column=col, value=header.replace('_', ' ').title())
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment

                current_row += 1

                # Data rows
                for record in records:
                    for col, header in enumerate(headers, 1):
                        value = record.get(header, '')
                        if value is None:
                            value = ''
                        ws.cell(row=current_row, column=col, value=value)
                    current_row += 1

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Save to bytes
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    def export_to_csv(self, data: Dict[str, Any], report_type: str) -> bytes:
        """Export report data to CSV"""
        if 'data' not in data or not data['data']:
            return b"No data available\n"

        if PANDAS_AVAILABLE:
            # Use pandas if available
            df = pd.DataFrame(data['data'])
            buffer = io.StringIO()
            df.to_csv(buffer, index=False, encoding='utf-8')
            buffer.seek(0)
            return buffer.getvalue().encode('utf-8')
        else:
            # Fallback to manual CSV generation
            import csv
            output = io.StringIO()

            if data['data']:
                fieldnames = data['data'][0].keys()
                writer = csv.DictWriter(output, fieldnames=fieldnames)
                writer.writeheader()
                for row in data['data']:
                    writer.writerow(row)

            output.seek(0)
            return output.getvalue().encode('utf-8')